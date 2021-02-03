#!/usr/bin/python3
# date 2021-02-03
# version 2.0
# add 1:结果说明sheet页生成
# add 2:循环对比多个sheet页

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, colors, Alignment


# 指定文件
file1 = "./测试数据.xlsx"  # 对比文件1
file2 = "./测试数据2.xlsx"  # 对比文件1
file3 = "./对比结果.xlsx"  # 结果文件

# 打开文件
work_book1 = openpyxl.load_workbook(file1)
work_book2 = openpyxl.load_workbook(file2)
# 准备结果文件
work_book3 = openpyxl.Workbook()
# 获取sheet页名称
sheet_name_list1 = work_book1.sheetnames
sheet_name_list2 = work_book2.sheetnames

# 判断是否有相同sheet页
same_sheet_list = []  # 定义一个列表用于存放相同sheet页
for sheet_name in sheet_name_list2:
    if sheet_name in sheet_name_list1:
        same_sheet_list.append(sheet_name)


# 设置不同数据样式，这里是红色字体
bold_itatic_24_font = Font(color='FF0000')


# 根据相同sheet名称列表中的sheet名称，循环对比sheet页
for sheet_name in same_sheet_list:
    print(sheet_name + "对比开始")
    work_sheet1 = work_book1[sheet_name]
    work_sheet2 = work_book2[sheet_name]
    work_sheet3 = work_book3.create_sheet(sheet_name, 0)

    max_row = max(work_sheet1.max_row, work_sheet2.max_row)
    max_col = max(work_sheet1.max_column, work_sheet2.max_column)

    # 遍历所有行
    for x in range(1, max_row+1):
        # 遍历所有列
        for y in range(1, max_col+1):
            # 获取当前列的英文字母
            colNum = get_column_letter(y)
            # 获取行号
            rowNum = x
            # 得到每行每列的所有数据，逐一对比
            if work_sheet1[colNum + str(rowNum)].value == work_sheet2[colNum + str(rowNum)].value:
                # 如果相等，结果文件里写入文件1的数据
                work_sheet3[colNum +
                            str(rowNum)] = work_sheet2[colNum + str(rowNum)].value
            else:
                # 如果不相等，结果文件里写入文件2的数据
                work_sheet3[colNum +
                            str(rowNum)] = work_sheet2[colNum + str(rowNum)].value
                # 并在结果文件中使用特殊样式进行标记
                work_sheet3[colNum + str(rowNum)].font = bold_itatic_24_font

    print(sheet_name + "对比结束")

# 处理两个文件中sheet名称不同的sheet页
work_sheet3 = work_book3.create_sheet("对比结果", 0)

work_sheet3["A1"] = "本文的其余sheet页中，黑色字体数据为两个文件相同的数据"
work_sheet3["B1"] = "红色字体数据为两个文件不同的数据。记录的数据是" + file2 + "的数据"

work_sheet3["A2"] = file1 + "中存在而" + file2 + "中不存在的sheet："
work_sheet3["A3"] = str(set(sheet_name_list1).difference(
    set(same_sheet_list)))  # sheet_name_list1 - same_sheet_list的差集

work_sheet3["A4"] = file2 + "中存在而" + file1 + "中不存在的sheet："
work_sheet3["A5"] = str(set(sheet_name_list2).difference(
    set(same_sheet_list)))  # sheet_name_list2 - same_sheet_list的差集

work_book3.save(file3)
print("数据写入：" + file3)